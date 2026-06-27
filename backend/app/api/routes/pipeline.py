"""
Deal Pipeline API Routes
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Path as PathParam, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List, Optional
from decimal import Decimal
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, func, or_, cast, String
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timezone
UTC = timezone.utc
import uuid
import io
import re

from app.core.database import get_db
from app.api.deps import get_current_user, require_facilitator, require_admin, has_twg_access
from app.models.models import User, Project, ProjectStatus, UserRole, ProjectInterest

# Roles allowed to see INCUBATION-status projects. All other roles (TWG_MEMBER,
# any future investor role) get incubation projects filtered out at the API.
INCUBATION_VISIBLE_ROLES = {UserRole.ADMIN, UserRole.SECRETARIAT_LEAD, UserRole.TWG_FACILITATOR}
from app.services.project_pipeline_service import ProjectPipelineService
from app.services.investor_matching_service import get_investor_matching_service
from app.services.dfi_matching_service import get_dfi_matching_service
from app.schemas.pipeline_schemas import (
    ProjectIngest, ProjectUpdate, ProjectPipelineRead, ProjectAdvanceStage,
    InvestorMatchRead, PipelineStats, InvestorMatchUpdate, InvestorRead,
    ProjectScoreDetailRead, ScoringCriteriaRead, ScoringCriteriaWeightUpdate,
    ReadinessGapRead, ReadinessGapItem,
    BuyerCreate, BuyerRead, BuyerMatchRead, BuyerMatchUpdate,
    DFIWindowRead, DFIMatchRead, DFIMatchStatusUpdate, FinancingMemoResponse,
    IncubationChecklistRead, IncubationChecklistItem,
    ProjectGeospatialRead, ImpactLogEntryCreate, ImpactLogEntryRead, ImpactSummaryRead,
    ProjectMemberRead, ProjectMemberDetail, ScoreBreakdownItem, ProjectInterestState,
)
from app.models.models import Document, ImpactLogEntry, ProjectGeospatialData
from app.core.constants import INCUBATION_CHECKLIST_ITEMS, canonical_code_for
from app.models.models import ProjectScoreDetail, ScoringCriteria, Buyer, DFIWindow, ProjectDFIMatch, DFIMatchStatus
from app.services.geospatial_service import get_geospatial_service
from app.services.coordinate_scout_service import get_coordinate_scout_service
from app.services.lifecycle_service import LifecycleService
from app.services.project_insights_service import insights_service

router = APIRouter(prefix="/pipeline", tags=["deal-pipeline"])


def _project_to_read(p: "Project", current_user: "User") -> ProjectPipelineRead:
    """Convert a Project ORM object to ProjectPipelineRead schema."""
    return ProjectPipelineRead(
        id=p.id,
        name=p.name,
        description=p.description,
        status=p.status,
        investment_size=p.investment_size,
        currency=p.currency,
        readiness_score=p.readiness_score,
        afcen_score=p.afcen_score,
        strategic_alignment_score=p.strategic_alignment_score,
        lead_country=p.lead_country,
        pillar=p.pillar,
        assigned_agent=p.assigned_agent,
        updated_at=getattr(p, 'created_at', datetime.now(UTC)),
        is_flagship=p.is_flagship,
        funding_secured_usd=p.funding_secured_usd or 0,
        deal_room_priority=p.deal_room_priority,
        # Section A
        subsector=p.subsector,
        project_sponsor=p.project_sponsor,
        is_cross_border=p.is_cross_border or False,
        key_contact_name=p.key_contact_name,
        key_contact_email=p.key_contact_email,
        # Section B
        technical_studies=p.technical_studies,
        permits_licences=p.permits_licences,
        land_status=p.land_status,
        # Section C
        financing_structure=p.financing_structure,
        investment_stage_label=p.investment_stage_label,
        revenue_model=p.revenue_model,
        macroeconomic_roi=p.macroeconomic_roi,
        # Section D
        climate_impact=p.climate_impact,
        esg_compliance=p.esg_compliance,
        ghg_avoided_target=p.ghg_avoided_target,
        jobs_construction=p.jobs_construction,
        jobs_om=p.jobs_om,
        electricity_connections=p.electricity_connections,
        digital_connections=p.digital_connections,
        smallholder_farmers_reached=p.smallholder_farmers_reached,
        submitted_by=p.submitted_by,
        # Phase 1 — Classification fields
        value_chain_stages=p.value_chain_stages,
        women_employment_pct=p.women_employment_pct,
        youth_employment_pct=p.youth_employment_pct,
        # R2 — Gender & Youth intentional design flags
        gender_intentional=p.gender_intentional,
        gender_justification=p.gender_justification,
        youth_focused=p.youth_focused,
        youth_justification=p.youth_justification,
        # R8 — Site coordinates
        site_lat=p.site_lat,
        site_lon=p.site_lon,
        site_location_name=p.site_location_name,
        allowed_transitions=LifecycleService.get_allowed_transitions(p.status, current_user.role),
    )


@router.get("/", response_model=List[ProjectPipelineRead])
async def list_pipeline_projects(
    stage: Optional[ProjectStatus] = None,
    pillar: Optional[str] = None,
    value_chain_stage: Optional[str] = None,
    include_archived: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    List projects in the deal pipeline with optional filtering.
    value_chain_stage: filter to projects that include this stage, e.g. 'INPUTS'
    include_archived: when false (default) ARCHIVED projects are omitted; admins
    can pass include_archived=true to see them.
    """
    query = select(Project)

    if stage:
        query = query.where(Project.status == stage)
    if pillar:
        # Case-insensitive pillar filter. Also match cross-listed projects whose
        # metadata_json.cross_listed_pillars contains this pillar, so a project
        # can appear under more than one pillar (e.g. digital-for-minerals systems
        # surfaced under both Digital Transformation and Strategic Minerals).
        query = query.where(or_(
            Project.pillar.ilike(f"%{pillar}%"),
            cast(Project.metadata_json, String).ilike(f"%{pillar}%"),
        ))
    if value_chain_stage:
        query = query.where(Project.value_chain_stages.contains([value_chain_stage]))

    # Hide INCUBATION projects from non-privileged roles (investors, TWG members)
    if current_user.role not in INCUBATION_VISIBLE_ROLES:
        query = query.where(Project.status != ProjectStatus.INCUBATION)

    # Hide ARCHIVED projects by default (mirrors the showIncubation pattern)
    if not include_archived and stage != ProjectStatus.ARCHIVED:
        query = query.where(Project.status != ProjectStatus.ARCHIVED)

    result = await db.execute(query)
    projects = result.scalars().all()

    return [_project_to_read(p, current_user) for p in projects]


# ---------------------------------------------------------------------------
# Member Deal Room — TWG-scoped read + interest (follow) toggle.
# NOTE: /member MUST be registered before /{project_id} (UUID path) below.
# ---------------------------------------------------------------------------

def _member_pillar_values(user: "User") -> List[str]:
    """DISTINCT pillar value strings of the user's TWGs.

    TWG.pillar is an Enum(TWGPillar) whose .value strings match Project.pillar
    strings (e.g. 'agriculture_food_systems') — compare on the value, and stay
    safe whether the ORM hands back the enum or a bare string.
    """
    values: List[str] = []
    for twg in user.twgs:
        pillar = getattr(twg, "pillar", None)
        if pillar is None:
            continue
        value = pillar.value if hasattr(pillar, "value") else str(pillar)
        if value not in values:
            values.append(value)
    return values


def _member_can_access_project(user: "User", project: "Project") -> bool:
    """Member visibility = TWG link OR TWG-pillar match.

    PROD DATA REALITY: projects are systematically linked to the wrong TWG row
    (e.g. agriculture projects carrying the Energy TWG's twg_id), so the twg_id
    link alone hides deals members can already see on the staff web pipeline.
    """
    if has_twg_access(user, project.twg_id):
        return True
    return project.pillar is not None and project.pillar in _member_pillar_values(user)


def _project_to_member_read(p: "Project", interest_count: int, is_following: bool) -> ProjectMemberRead:
    """Member-safe projection — NO key contacts / financing internals."""
    return ProjectMemberRead(
        id=p.id,
        twg_id=p.twg_id,
        name=p.name,
        sector=p.pillar,
        status=p.status,
        investment_size=p.investment_size,
        currency=p.currency or "USD",
        readiness_score=p.readiness_score or 0.0,
        afcen_score=p.afcen_score,
        strategic_alignment_score=p.strategic_alignment_score,
        location=p.lead_country or p.site_location_name,
        description=p.description,
        is_following=is_following,
        interest_count=interest_count,
    )


@router.get("/member", response_model=List[ProjectMemberRead])
async def list_member_projects(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Member-facing Deal Room list: projects linked to the caller's TWG(s).

    Unlike the staff pipeline list above, this shows ALL lifecycle stages —
    including INCUBATION — because members may see their OWN TWG's early
    projects; what they may never see is other TWGs' projects. ARCHIVED is
    excluded. Serialization is member-safe (ProjectMemberRead) only.
    """
    query = select(Project).where(Project.status != ProjectStatus.ARCHIVED)

    # Admin / Secretariat see everything; everyone else: TWG link OR TWG pillar.
    # PROD DATA REALITY: projects are systematically linked to the wrong TWG row,
    # so twg_id alone empties the member Deal Room for deals the staff pipeline
    # (pillar-based, no TWG restriction) already shows on the web.
    if current_user.role not in (UserRole.ADMIN, UserRole.SECRETARIAT_LEAD):
        twg_ids = [twg.id for twg in current_user.twgs]
        if not twg_ids:
            return []
        access_clause = Project.twg_id.in_(twg_ids)
        pillar_values = _member_pillar_values(current_user)
        if pillar_values:
            access_clause = or_(access_clause, Project.pillar.in_(pillar_values))
        query = query.where(access_clause)

    result = await db.execute(query.order_by(Project.name))
    projects = result.scalars().all()
    if not projects:
        return []

    project_ids = [p.id for p in projects]
    count_rows = await db.execute(
        select(ProjectInterest.project_id, func.count(ProjectInterest.id))
        .where(ProjectInterest.project_id.in_(project_ids))
        .group_by(ProjectInterest.project_id)
    )
    counts = dict(count_rows.all())
    mine_rows = await db.execute(
        select(ProjectInterest.project_id).where(
            ProjectInterest.user_id == current_user.id,
            ProjectInterest.project_id.in_(project_ids),
        )
    )
    mine = set(mine_rows.scalars().all())

    return [_project_to_member_read(p, counts.get(p.id, 0), p.id in mine) for p in projects]


async def _get_member_project_or_404(
    project_id: uuid.UUID, db: AsyncSession, current_user: User
) -> Project:
    """Load a project the caller's TWGs grant access to (TWG link OR TWG pillar
    — the same visibility rule as the /member list).

    Cross-TWG (and nonexistent) projects both return 404 — not 403 — so
    members cannot enumerate other TWGs' deal flow.
    """
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project or not _member_can_access_project(current_user, project):
        raise HTTPException(status_code=404, detail="Project not found")
    return project


async def _interest_count(db: AsyncSession, project_id: uuid.UUID) -> int:
    result = await db.execute(
        select(func.count()).select_from(ProjectInterest)
        .where(ProjectInterest.project_id == project_id)
    )
    return result.scalar_one()


@router.get("/member/{project_id}", response_model=ProjectMemberDetail)
async def get_member_project_detail(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Member-facing Deal Room DETAIL: the member-safe list projection plus
    descriptive extras and a stripped score breakdown.

    Same visibility rule as the /member list (TWG link OR TWG pillar; 404 —
    never 403 — on cross-TWG/nonexistent, so deal flow can't be enumerated).
    ARCHIVED projects 404 here too, mirroring their exclusion from the list.
    The breakdown carries criterion/weight/score ONLY — scorer notes and
    identities are facilitator-side and never serialized for members.
    """
    project = await _get_member_project_or_404(project_id, db, current_user)
    if project.status == ProjectStatus.ARCHIVED:
        raise HTTPException(status_code=404, detail="Project not found")

    is_following = (await db.execute(
        select(ProjectInterest.id).where(
            ProjectInterest.project_id == project.id,
            ProjectInterest.user_id == current_user.id,
        )
    )).scalar_one_or_none() is not None

    score_rows = await db.execute(
        select(ScoringCriteria.criterion_name, ScoringCriteria.weight, ProjectScoreDetail.score)
        .join(ScoringCriteria, ProjectScoreDetail.criterion_id == ScoringCriteria.id)
        .where(ProjectScoreDetail.project_id == project.id)
        .order_by(desc(ScoringCriteria.weight))
    )
    score_breakdown = [
        ScoreBreakdownItem(criterion=name, weight=float(weight), score=float(score))
        for name, weight, score in score_rows.all()
    ]

    base = _project_to_member_read(
        project, await _interest_count(db, project.id), is_following
    )
    return ProjectMemberDetail(
        **base.model_dump(),
        subsector=project.subsector,
        investment_stage_label=project.investment_stage_label,
        project_sponsor=project.project_sponsor,
        is_cross_border=project.is_cross_border,
        financing_structure=project.financing_structure,
        technical_studies=project.technical_studies,
        land_status=project.land_status,
        permits_licences=project.permits_licences,
        climate_impact=project.climate_impact,
        smallholder_farmers_reached=project.smallholder_farmers_reached,
        submitted_by=project.submitted_by,
        updated_at=project.updated_at,
        score_breakdown=score_breakdown,
    )


@router.post("/{project_id}/interest", response_model=ProjectInterestState)
async def express_project_interest(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Follow / express interest in a project (idempotent). TWG-checked."""
    project = await _get_member_project_or_404(project_id, db, current_user)

    existing = (await db.execute(
        select(ProjectInterest).where(
            ProjectInterest.project_id == project.id,
            ProjectInterest.user_id == current_user.id,
        )
    )).scalar_one_or_none()
    if not existing:
        db.add(ProjectInterest(project_id=project.id, user_id=current_user.id))
        try:
            await db.commit()
        except IntegrityError:
            # Concurrent double-tap — the unique constraint already holds the row.
            await db.rollback()

    return ProjectInterestState(
        project_id=project.id,
        is_following=True,
        interest_count=await _interest_count(db, project.id),
    )


@router.delete("/{project_id}/interest", response_model=ProjectInterestState)
async def remove_project_interest(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Unfollow a project (idempotent). TWG-checked."""
    project = await _get_member_project_or_404(project_id, db, current_user)

    existing = (await db.execute(
        select(ProjectInterest).where(
            ProjectInterest.project_id == project.id,
            ProjectInterest.user_id == current_user.id,
        )
    )).scalar_one_or_none()
    if existing:
        await db.delete(existing)
        await db.commit()

    return ProjectInterestState(
        project_id=project.id,
        is_following=False,
        interest_count=await _interest_count(db, project.id),
    )


@router.post("/ingest", response_model=ProjectPipelineRead)
async def ingest_project(
    data: ProjectIngest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    """
    Ingest a new project proposal and calculate initial scores.
    """
    service = ProjectPipelineService(db)
    
    result = await service.ingest_project_proposal(
        data=data.model_dump(exclude={"start_in_incubation"}),
        submitted_by_user_id=current_user.id,
        start_in_incubation=data.start_in_incubation,
    )
    
    p = result["project"]
    return _project_to_read(p, current_user)


@router.get("/scoring-criteria", response_model=List[ScoringCriteriaRead])
async def list_scoring_criteria(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all WAIIS scoring criteria and their current weights."""
    result = await db.execute(select(ScoringCriteria))
    criteria = result.scalars().all()
    return [
        ScoringCriteriaRead(
            id=c.id,
            criterion_name=c.criterion_name,
            criterion_type=c.criterion_type,
            weight=c.weight,
            description=c.description
        )
        for c in criteria
    ]


@router.patch("/scoring-criteria/{criterion_id}", response_model=ScoringCriteriaRead)
async def update_scoring_criterion_weight(
    criterion_id: uuid.UUID,
    body: ScoringCriteriaWeightUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Update the weight of a WAIIS scoring criterion. Admin only."""
    from decimal import Decimal
    result = await db.execute(select(ScoringCriteria).where(ScoringCriteria.id == criterion_id))
    criterion = result.scalar_one_or_none()
    if not criterion:
        raise HTTPException(status_code=404, detail="Criterion not found")
    if body.weight < Decimal("0") or body.weight > Decimal("9.99"):
        raise HTTPException(status_code=422, detail="Weight must be between 0 and 9.99")
    criterion.weight = body.weight
    await db.commit()
    return ScoringCriteriaRead(
        id=criterion.id,
        criterion_name=criterion.criterion_name,
        criterion_type=criterion.criterion_type,
        weight=criterion.weight,
        description=criterion.description
    )


# ---------------------------------------------------------------------------
# Platform Settings endpoints — MUST be before /{project_id} to avoid routing clash
# ---------------------------------------------------------------------------

@router.get("/settings")
async def get_platform_settings(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Return all platform settings as key/value pairs. Admin only."""
    from app.models.models import PlatformSetting
    result = await db.execute(select(PlatformSetting))
    settings = result.scalars().all()
    return {s.key: s.value for s in settings}


@router.patch("/settings")
async def update_platform_settings(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Update one or more platform settings. Admin only."""
    from app.models.models import PlatformSetting
    ALLOWED_KEYS = {"gender_threshold_pct", "youth_threshold_pct", "incubation_graduation_threshold"}
    for key, value in payload.items():
        if key not in ALLOWED_KEYS:
            raise HTTPException(status_code=400, detail=f"Unknown setting key: {key}")
        result = await db.execute(select(PlatformSetting).where(PlatformSetting.key == key))
        setting = result.scalars().first()
        if setting:
            setting.value = str(value)
        else:
            db.add(PlatformSetting(key=key, value=str(value)))
    await db.commit()
    return {"updated": list(payload.keys())}


# ---------------------------------------------------------------------------
# Buyer CRUD + buyer-match PATCH — MUST be before /{project_id} to avoid UUID routing clash
# ---------------------------------------------------------------------------
from app.services.buyer_matching_service import get_buyer_matching_service


@router.get("/buyers/", response_model=List[BuyerRead])
async def list_buyers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = await db.execute(select(Buyer).where(Buyer.deleted_at.is_(None)))
    buyers = result.scalars().all()
    return [BuyerRead.model_validate(b) for b in buyers]


@router.post("/buyers/", response_model=BuyerRead, status_code=201)
async def create_buyer(
    payload: BuyerCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    buyer = Buyer(
        id=uuid.uuid4(),
        name=payload.name,
        commodity_types=payload.commodity_types,
        volume_mt_per_year=payload.volume_mt_per_year,
        contract_term_years=payload.contract_term_years,
        price_floor_usd=payload.price_floor_usd,
        geographic_focus=payload.geographic_focus,
        notes=payload.notes,
        created_by=current_user.id,
    )
    db.add(buyer)
    await db.commit()
    await db.refresh(buyer)
    return BuyerRead.model_validate(buyer)


@router.patch("/buyer-matches/{match_id}", response_model=dict)
async def update_buyer_match_status(
    match_id: uuid.UUID,
    payload: BuyerMatchUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    service = get_buyer_matching_service(db)
    result = await service.update_match_status(
        match_id=match_id,
        new_status=payload.status,
        notes=payload.notes,
    )
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


# ── DFI Windows ───────────────────────────────────────────────────────────────

@router.get("/dfi-windows", response_model=List[DFIWindowRead])
async def list_dfi_windows(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all active DFI/climate-finance windows."""
    result = await db.execute(
        select(DFIWindow).where(DFIWindow.is_active.is_(True)).order_by(DFIWindow.institution)
    )
    return result.scalars().all()


@router.patch("/dfi-matches/{match_id}", response_model=dict)
async def update_dfi_match_status(
    match_id: uuid.UUID,
    payload: DFIMatchStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update the status (and optional notes) of a DFI window match."""
    svc = get_dfi_matching_service(db)
    result = await svc.update_match_status(match_id, payload.status, payload.notes)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


# ---------------------------------------------------------------------------

@router.get("/templates/financial-model")
async def download_financial_model_template(
    current_user: User = Depends(get_current_user),
):
    """Return a generated WAIIS financial-model XLSX template.

    Any authenticated user may download. The template includes 7 sheets covering
    the WAIIS-required line items (Sources & Uses, 5Y P&L, Capex, Sensitivity,
    Currency Exposure, ESIA Costs, Social Impact). Sheets are blank — sponsors
    fill in their own numbers.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2D3748")
        instr_font = Font(italic=True, color="555555", size=10)
        section_font = Font(bold=True, size=11)

        def write_sheet(ws, title: str, instruction: str, headers: list[str], rows: list[list[str]]):
            ws.title = title
            ws["A1"] = instruction
            ws["A1"].font = instr_font
            ws.row_dimensions[1].height = 22
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
            for r_idx, row in enumerate(rows, start=4):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            # column widths
            for c in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + c)].width = 22

        # --- Sheet 1: Sources & Uses ---
        ws1 = wb.active
        write_sheet(
            ws1,
            "Sources & Uses",
            "Capital stack. Leave rows that don't apply blank. Sum of Sources must equal Uses.",
            ["Source / Use", "Amount (USD)", "Tranche / Window", "Notes"],
            [
                ["SOURCES", "", "", ""],
                ["DFI tranche 1", "", "", ""],
                ["DFI tranche 2", "", "", ""],
                ["Sponsor equity", "", "", ""],
                ["Grant funding", "", "", ""],
                ["Commercial debt", "", "", ""],
                ["Other concessional", "", "", ""],
                ["", "", "", ""],
                ["USES", "", "", ""],
                ["Land acquisition", "", "", ""],
                ["Construction / CapEx", "", "", ""],
                ["Equipment", "", "", ""],
                ["Working capital", "", "", ""],
                ["Contingency", "", "", ""],
            ],
        )

        # --- Sheet 2: 5Y P&L ---
        ws2 = wb.create_sheet()
        write_sheet(
            ws2,
            "5Y P&L",
            "5-year projection. Enter values in USD, in nominal terms. Year 1 = first full operating year.",
            ["Line Item", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"],
            [
                ["Revenue", "", "", "", "", ""],
                ["COGS", "", "", "", "", ""],
                ["Gross Profit", "", "", "", "", ""],
                ["Operating Expenses", "", "", "", "", ""],
                ["EBITDA", "", "", "", "", ""],
                ["Depreciation", "", "", "", "", ""],
                ["EBIT", "", "", "", "", ""],
                ["Interest expense", "", "", "", "", ""],
                ["Taxes", "", "", "", "", ""],
                ["Net Income", "", "", "", "", ""],
            ],
        )

        # --- Sheet 3: Capex Schedule ---
        ws3 = wb.create_sheet()
        write_sheet(
            ws3,
            "Capex Schedule",
            "Construction-phase capital expenditure by quarter. Add rows as needed for line items.",
            ["Capex Line Item", "Q1", "Q2", "Q3", "Q4", "Total"],
            [
                ["Site preparation", "", "", "", "", ""],
                ["Civil works", "", "", "", "", ""],
                ["Equipment procurement", "", "", "", "", ""],
                ["Installation & commissioning", "", "", "", "", ""],
                ["Owner's costs", "", "", "", "", ""],
                ["Contingency", "", "", "", "", ""],
            ],
        )

        # --- Sheet 4: Sensitivity ---
        ws4 = wb.create_sheet()
        write_sheet(
            ws4,
            "Sensitivity",
            "IRR / NPV at three commodity-price scenarios. Document each scenario's assumption.",
            ["Scenario", "Commodity price assumption", "IRR (%)", "NPV (USD)"],
            [
                ["Low", "", "", ""],
                ["Base", "", "", ""],
                ["High", "", "", ""],
            ],
        )

        # --- Sheet 5: Currency Exposure ---
        ws5 = wb.create_sheet()
        write_sheet(
            ws5,
            "Currency Exposure",
            "Share of revenue and costs denominated in USD vs local currency.",
            ["Category", "USD %", "Local currency %", "Notes"],
            [
                ["Revenue", "", "", ""],
                ["Operating costs", "", "", ""],
                ["Capex", "", "", ""],
                ["Debt service", "", "", ""],
            ],
        )

        # --- Sheet 6: ESIA Costs ---
        ws6 = wb.create_sheet()
        write_sheet(
            ws6,
            "ESIA Costs",
            "Environmental & social compliance budget. Required for DFI eligibility.",
            ["ESIA Line Item", "Amount (USD)", "Phase", "Notes"],
            [
                ["Environmental impact assessment", "", "Pre-construction", ""],
                ["Social impact assessment", "", "Pre-construction", ""],
                ["Community consultation", "", "Pre-construction", ""],
                ["Resettlement / compensation", "", "Pre-construction", ""],
                ["Monitoring & reporting", "", "Operation", ""],
                ["Closure / rehabilitation", "", "End-of-life", ""],
            ],
        )

        # --- Sheet 7: Social Impact ---
        ws7 = wb.create_sheet()
        write_sheet(
            ws7,
            "Social Impact",
            "Baseline social-impact metrics. These feed WAIIS scoring directly.",
            ["Metric", "Target value", "Methodology / notes"],
            [
                ["Jobs created — construction phase", "", ""],
                ["Jobs created — O&M phase", "", ""],
                ["Smallholders reached", "", ""],
                ["Women employment %", "", ""],
                ["Youth employment %", "", ""],
                ["Electricity connections enabled", "", ""],
                ["Digital connections enabled", "", ""],
            ],
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="waiis_financial_model_template.xlsx"',
            },
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Template generation failed: {exc}")


@router.get("/{project_id}", response_model=ProjectPipelineRead)
async def get_project_details(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get detailed project view.
    """
    result = await db.execute(select(Project).where(Project.id == project_id))
    p = result.scalars().first()

    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    # Hide INCUBATION projects from non-privileged roles — return 404 (not 403)
    # to avoid leaking the existence of incubation projects via enumeration.
    if p.status == ProjectStatus.INCUBATION and current_user.role not in INCUBATION_VISIBLE_ROLES:
        raise HTTPException(status_code=404, detail="Project not found")

    return _project_to_read(p, current_user)

@router.patch("/{project_id}", response_model=ProjectPipelineRead)
async def update_project(
    project_id: uuid.UUID,
    payload: ProjectUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    """
    Update project details.
    """
    service = ProjectPipelineService(db)
    result = await service.update_project(
        project_id=project_id,
        data=payload.model_dump(exclude_unset=True),
        updated_by_user_id=current_user.id
    )

    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])

    p = result["project"]
    return _project_to_read(p, current_user)

@router.post("/{project_id}/advance", response_model=ProjectPipelineRead)
async def advance_project_stage(
    project_id: uuid.UUID,
    payload: ProjectAdvanceStage,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    """
    Advance a project to the next stage.
    """
    service = ProjectPipelineService(db)
    
    result = await service.advance_project_stage(
        project_id=project_id,
        new_stage=payload.new_stage,
        advanced_by_user_id=current_user.id,
        notes=payload.notes
    )
    
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
        
    p = result["project"]
    return _project_to_read(p, current_user)

@router.get("/{project_id}/matches", response_model=List[InvestorMatchRead])
async def get_project_matches(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get investor matches for a project.
    """
    service = get_investor_matching_service(db)
    matches = await service.get_matches_for_project(project_id)
    
    return [
        InvestorMatchRead(
            match_id=m["match_id"],
            investor=InvestorRead(
                id=m["investor"].id,
                name=m["investor"].name,
                sector_preferences=m["investor"].sector_preferences,
                ticket_size_min=m["investor"].ticket_size_min,
                ticket_size_max=m["investor"].ticket_size_max,
                geographic_focus=m["investor"].geographic_focus,
                investment_instruments=m["investor"].investment_instruments
            ),
            score=m["score"],
            status=m["status"],
            notes=m["notes"]
        ) for m in matches
    ]

@router.patch("/matches/{match_id}", response_model=dict)
async def update_match_status(
    match_id: uuid.UUID,
    payload: InvestorMatchUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    """
    Update match status (e.g. to INTERESTED to trigger Protocol Agent).
    """
    service = get_investor_matching_service(db)
    result = await service.update_match_status(
        match_id=match_id,
        new_status=payload.status,
        notes=payload.notes,
        updated_by_user_id=current_user.id
    )
    
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
        
    return result

@router.post("/{project_id}/match", response_model=dict)
async def trigger_investor_matching(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    """
    Manually trigger investor matching for a project.
    """
    service = get_investor_matching_service(db)
    result = await service.match_investors(project_id)
    return result

@router.get("/dashboard/stats", response_model=PipelineStats)
async def get_pipeline_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get high-level pipeline statistics.
    """
    service = ProjectPipelineService(db)
    stats = await service.check_pipeline_health()
    return PipelineStats(**stats)

@router.get("/{project_id}/score-details", response_model=List[ProjectScoreDetailRead])
async def get_project_score_details(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get detailed breakdown of project scores.
    """
    # Join with Criteria
    stmt = (
        select(ProjectScoreDetail, ScoringCriteria)
        .join(ScoringCriteria, ProjectScoreDetail.criterion_id == ScoringCriteria.id)
        .where(ProjectScoreDetail.project_id == project_id)
    )
    result = await db.execute(stmt)
    rows = result.all()
    
    return [
        ProjectScoreDetailRead(
            id=d.id,
            criterion=ScoringCriteriaRead(
                id=c.id,
                criterion_name=c.criterion_name,
                criterion_type=c.criterion_type,
                weight=c.weight,
                description=c.description
            ),
            score=float(d.score),
            notes=d.notes,
            scored_date=d.scored_date
        )
        for d, c in rows
    ]


@router.get("/{project_id}/insights")
async def get_project_insights(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get AI-powered insights and recommendations for a project.
    Uses LLM to analyze project status, scores, and generate actionable next steps.
    """
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    insights = await insights_service.generate_project_insights(project, db)
    
    return {
        "project_id": str(project_id),
        "insight": insights["insight"],
        "recommendation": insights["recommendation"],
        "generated_at": datetime.now(UTC).isoformat()
    }


@router.get("/{project_id}/history")
async def get_project_history(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get project status change history timeline.
    Returns chronological list of all status changes.
    """
    from app.models.models import ProjectStatusHistory, User as UserModel
    
    # Verify project exists
    project_result = await db.execute(select(Project).where(Project.id == project_id))
    project = project_result.scalar_one_or_none()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Fetch status history with user info
    stmt = (
        select(ProjectStatusHistory, UserModel)
        .outerjoin(UserModel, ProjectStatusHistory.changed_by_id == UserModel.id)
        .where(ProjectStatusHistory.project_id == project_id)
        .order_by(ProjectStatusHistory.change_date.desc())
    )
    
    result = await db.execute(stmt)
    history_records = result.all()
    
    # Format response
    history = []
    for record, user in history_records:
        history.append({
            "id": str(record.id),
            "previous_status": record.previous_status.value if record.previous_status else None,
            "new_status": record.new_status.value,
            "change_date": record.change_date.isoformat(),
            "changed_by": {
                "id": str(user.id) if user else None,
                "name": user.full_name if user else "System",
                "email": user.email if user else None
            } if user else {"name": "System"},
            "notes": record.notes
        })
    
    return {
        "project_id": str(project_id),
        "history": history,
        "total_changes": len(history)
    }


@router.post("/{project_id}/feature")
async def toggle_project_flagship(
    project_id: uuid.UUID,
    is_flagship: bool = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    """
    Toggle the 'is_flagship' status of a project.
    """
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalars().first()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
        
    project.is_flagship = is_flagship
    await db.commit()
    
    return {"status": "success", "is_flagship": is_flagship}


@router.post("/{project_id}/rescore")
async def rescore_project(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    """
    Manually trigger AfCEN scoring assessment for a project.
    Useful when automatic Celery scoring is unavailable.
    """
    # Verify project exists
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Run scoring synchronously
    service = ProjectPipelineService(db)
    try:
        score = await service.assess_project_readiness(project_id)

        return {
            "status": "success",
            "project_id": str(project_id),
            "afcen_score": float(score),
            "message": f"Project rescored successfully. New AfCEN score: {score:.2f}"
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Scoring failed: {str(e)}"
        )


@router.get("/{project_id}/readiness-gap", response_model=ReadinessGapRead)
async def get_readiness_gap(
    project_id: uuid.UUID,
    refresh: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator),
):
    """Generate (or return cached) Martin gap report for an Incubation project."""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if project.status != ProjectStatus.INCUBATION:
        raise HTTPException(status_code=400, detail="Readiness gap report is only available for Incubation projects")

    from app.models.models import PlatformSetting
    thresh_res = await db.execute(
        select(PlatformSetting).where(PlatformSetting.key == "incubation_graduation_threshold")
    )
    thresh_setting = thresh_res.scalars().first()
    threshold = float(thresh_setting.value) if thresh_setting else 40.0
    current_score = float(project.afcen_score or 0)

    # Return cached report if available (skip if refresh=true)
    meta = project.metadata_json or {}
    cached_report = meta.get("readiness_gap_report")
    if cached_report and not refresh:
        return ReadinessGapRead(
            gaps=[ReadinessGapItem(**g) for g in cached_report["gaps"]],
            current_score=current_score,
            threshold=threshold,
            cached=True,
        )

    # Fetch per-criterion scores
    from app.models.models import ProjectScoreDetail as _PSD, ScoringCriteria as _SC
    score_rows_result = await db.execute(
        select(_PSD, _SC)
        .join(_SC, _PSD.criterion_id == _SC.id)
        .where(_PSD.project_id == project_id)
    )
    criterion_scores = {
        row[1].criterion_name: {
            "score": float(row[0].score),
            "weight_pct": float(row[1].weight) * 10,
            "notes": row[0].notes or "",
        }
        for row in score_rows_result
    }

    project_fields = {
        "name": project.name,
        "description": project.description,
        "investment_size": str(project.investment_size),
        "lead_country": project.lead_country,
        "pillar": project.pillar,
        "project_sponsor": project.project_sponsor,
        "key_contact_name": project.key_contact_name,
        "financing_structure": project.financing_structure,
        "revenue_model": project.revenue_model,
        "technical_studies": project.technical_studies,
        "permits_licences": project.permits_licences,
        "land_status": project.land_status,
        "climate_impact": project.climate_impact,
        "esg_compliance": project.esg_compliance,
        "women_employment_pct": project.women_employment_pct,
        "youth_employment_pct": project.youth_employment_pct,
        "value_chain_stages": project.value_chain_stages,
        "is_cross_border": project.is_cross_border,
    }

    import json as _json
    prompt = (
        f"You are analysing an investment project for the ECOWAS Summit deal pipeline.\n"
        f"The project is in Incubation (pre-pipeline stage). Identify the 3-4 highest-impact gaps "
        f"preventing this project from reaching the graduation threshold of {threshold:.0f}/100.\n\n"
        f"Project data: {_json.dumps(project_fields, default=str)}\n"
        f"Current WAIIS scores per criterion: {_json.dumps(criterion_scores)}\n"
        f"Graduation threshold: {threshold:.0f}\nCurrent score: {current_score:.1f}\n\n"
        f"For each gap: criterion name, weight as percentage string (e.g. '18%'), "
        f"what is missing, and one concrete action referencing actual field names.\n"
        f"Output ONLY valid JSON: {{\"gaps\": [{{\"criterion\": \"...\", \"weight\": \"...\", \"issue\": \"...\", \"action\": \"...\"}}]}}"
    )

    from app.services.llm_service import llm_service
    gaps: list[ReadinessGapItem] = []
    try:
        raw = llm_service.chat(prompt, max_tokens=800)
        raw = raw.strip()
        if raw.startswith("```"):
            parts = raw.split("```")
            raw = parts[1] if len(parts) > 1 else raw
            if raw.startswith("json"):
                raw = raw[4:]
        parsed = _json.loads(raw.strip())
        gaps = [ReadinessGapItem(**g) for g in parsed["gaps"]]
    except Exception:
        for name, data in criterion_scores.items():
            if data["score"] == 0 and len(gaps) < 4:
                gaps.append(ReadinessGapItem(
                    criterion=name,
                    weight=f"{data['weight_pct']:.0f}%",
                    issue=f"{name} score is 0 — no relevant data provided",
                    action=f"Fill in {name.lower()} fields or upload supporting documents",
                ))

    # Cache
    meta["readiness_gap_report"] = {"gaps": [g.model_dump() for g in gaps]}
    project.metadata_json = meta
    await db.commit()

    return ReadinessGapRead(gaps=gaps, current_score=current_score, threshold=threshold, cached=False)


# ---------------------------------------------------------------------------
# Excel Import helpers
# ---------------------------------------------------------------------------

from app.models.models import TWGPillar, TWG


_PILLAR_KEYWORDS: dict[str, TWGPillar] = {
    "energy": TWGPillar.energy_infrastructure,
    "agriculture": TWGPillar.agriculture_food_systems,
    "agribusiness": TWGPillar.agriculture_food_systems,
    "food": TWGPillar.agriculture_food_systems,
    "mineral": TWGPillar.critical_minerals_industrialization,
    "industrial": TWGPillar.critical_minerals_industrialization,
    "digital": TWGPillar.digital_economy_transformation,
    "protocol": TWGPillar.protocol_logistics,
    "logistics": TWGPillar.protocol_logistics,
    "resource": TWGPillar.resource_mobilization,
    "mobilization": TWGPillar.resource_mobilization,
}

_STAGE_MAP: dict[str, ProjectStatus] = {
    "early-stage commercialisation": ProjectStatus.PIPELINE,
    "early-stage commercialization": ProjectStatus.PIPELINE,
    "feasibility / investment-ready": ProjectStatus.SUMMIT_READY,
    "feasibility / bankable": ProjectStatus.SUMMIT_READY,
    "pre-feasibility": ProjectStatus.PIPELINE,
    "prefeasibility": ProjectStatus.PIPELINE,
    "feasibility": ProjectStatus.UNDER_REVIEW,
    "bankable": ProjectStatus.SUMMIT_READY,
    "investment-ready": ProjectStatus.SUMMIT_READY,
    "investment ready": ProjectStatus.SUMMIT_READY,
    "concept": ProjectStatus.DRAFT,
    "draft": ProjectStatus.DRAFT,
    "early stage": ProjectStatus.PIPELINE,
}


def _match_pillar(raw: str) -> Optional[TWGPillar]:
    """Map a free-text sector string to a TWGPillar, or None if unrecognized.

    Returns None (NOT a Digital default) so the caller can fall back to the
    importing TWG's pillar. The old silent Digital default is what sent every
    Energy project from a ministry sheet into Digital Transformation.
    """
    lowered = (raw or "").strip().lower()
    if not lowered:
        return None
    for keyword, pillar in _PILLAR_KEYWORDS.items():
        if keyword in lowered:
            return pillar
    return None


def _map_status(raw: str) -> ProjectStatus:
    """Map a stage string to ProjectStatus; defaults to DRAFT."""
    lowered = raw.strip().lower()
    for key, status in _STAGE_MAP.items():
        if key in lowered:
            return status
    return ProjectStatus.DRAFT


def _parse_investment(raw: str) -> Optional[float]:
    """Parse investment value from strings like '$50M', '50–100 million', 'USD 100-150 million'."""
    if not raw:
        return None
    cleaned = str(raw).replace("$", "").replace(",", "").replace("USD", "").strip().lower()
    multiplier = 1_000_000 if any(x in cleaned for x in ("m", "million")) else 1
    # Handle ranges like "100–150" or "50-100" — take the lower bound
    range_match = re.search(r"(\d+(?:\.\d+)?)\s*[–\-]\s*(\d+(?:\.\d+)?)", cleaned)
    if range_match:
        low, high = float(range_match.group(1)), float(range_match.group(2))
        return ((low + high) / 2) * multiplier
    cleaned = re.sub(r"[^0-9.]", "", cleaned)
    try:
        return float(cleaned) * multiplier if cleaned else None
    except (ValueError, TypeError):
        return None


def _col_matches(header: str, *patterns: str) -> bool:
    """Case-insensitive partial match of a header against any of the patterns."""
    lowered = header.strip().lower()
    return any(p in lowered for p in patterns)


# A real header row matches at least this many known columns. A lone
# "Project Title" banner row matches 1, so this lets us skip banners/sections.
_MIN_HEADER_FIELDS = 3


def _build_col_map(header_cells_lower: list[str]) -> dict[str, int]:
    """Map known field names to 0-based column indices for one candidate row."""
    col_map: dict[str, int] = {}
    for col_idx, header in enumerate(header_cells_lower):
        if not header:
            continue
        if _col_matches(header, "project name", "project/programme name", "project title"):
            col_map.setdefault("name", col_idx)
        elif _col_matches(header, "country") and "lead country" not in header:
            col_map.setdefault("lead_country", col_idx)
        elif _col_matches(header, "cross-border", "national or cross-border", "cross border", "regional dimension"):
            col_map.setdefault("is_cross_border", col_idx)
        elif _col_matches(header, "sector") and "subsector" not in header:
            col_map.setdefault("pillar", col_idx)
        elif _col_matches(header, "subsector"):
            col_map.setdefault("subsector", col_idx)
        elif _col_matches(header, "project sponsor", "sponsor"):
            col_map.setdefault("project_sponsor", col_idx)
        elif _col_matches(header, "name of key contact", "key contact"):
            col_map.setdefault("key_contact_name", col_idx)
        elif _col_matches(header, "email of key contact", "email of key", "contact email"):
            col_map.setdefault("key_contact_email", col_idx)
        elif _col_matches(header, "stage of development"):
            col_map.setdefault("status", col_idx)
        elif _col_matches(header, "technical studies", "completed studies"):
            col_map.setdefault("technical_studies", col_idx)
        elif _col_matches(header, "permits", "licences", "licenses"):
            col_map.setdefault("permits_licences", col_idx)
        elif _col_matches(header, "land status", "land_status"):
            col_map.setdefault("land_status", col_idx)
        elif _col_matches(header, "investment size", "estimated investment", "capital required"):
            col_map.setdefault("investment_size", col_idx)
        elif _col_matches(header, "financing structure"):
            col_map.setdefault("financing_structure", col_idx)
        elif _col_matches(header, "investment stage"):
            col_map.setdefault("investment_stage_label", col_idx)
        elif _col_matches(header, "revenue model", "revenue_model"):
            col_map.setdefault("revenue_model", col_idx)
        elif _col_matches(header, "macroeconomic roi", "macro", "economic roi"):
            col_map.setdefault("macroeconomic_roi", col_idx)
        elif _col_matches(header, "climate", "esg"):
            col_map.setdefault("climate_impact", col_idx)
        elif _col_matches(header, "ghg", "tco2", "emissions"):
            col_map.setdefault("ghg_avoided_target", col_idx)
        elif _col_matches(header, "jobs", "construction") and "o&m" not in header and "ongoing" not in header:
            col_map.setdefault("jobs_construction", col_idx)
        elif _col_matches(header, "jobs", "o&m") or _col_matches(header, "jobs", "ongoing"):
            col_map.setdefault("jobs_om", col_idx)
        elif _col_matches(header, "electricity connect"):
            col_map.setdefault("electricity_connections", col_idx)
        elif _col_matches(header, "digital connect", "smes digitized"):
            col_map.setdefault("digital_connections", col_idx)
        elif _col_matches(header, "smallholder", "farmers reached"):
            col_map.setdefault("smallholder_farmers_reached", col_idx)
        elif _col_matches(header, "submitted by"):
            col_map.setdefault("submitted_by", col_idx)
    return col_map


def _find_header_row(ws):
    """
    Find the real header row by SCORING every row on how many known columns it
    matches, then picking the best (>= _MIN_HEADER_FIELDS). This is robust to
    ministry sheets that put a lone 'Project Title' banner (and Section A–D
    banners) above the actual header — those score 0–1 and are skipped.

    Returns (row_index_1based, col_map) or (None, None). If the winning header
    has no name cell (some sheets drop 'Project Title' from the real header row
    and float it into a banner), the name defaults to column 0 — the project
    title is the first column in every variant of this template.
    """
    best_idx = None
    best_map: dict[str, int] = {}
    best_score = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        col_map = _build_col_map(cells)
        if len(col_map) > best_score:
            best_score, best_idx, best_map = len(col_map), row_idx, col_map
    if best_idx is None or best_score < _MIN_HEADER_FIELDS:
        return None, None
    if "name" not in best_map:
        best_map["name"] = 0
    return best_idx, best_map


def _cell(row: tuple, col_map: dict, field: str) -> str:
    """Safely get a cell value as a stripped string, or empty string."""
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


# Names that are template scaffolding, not real projects.
_JUNK_NAME_MARKERS = (
    "full name of the project", "add further rows", "↓", "copy formatting",
    "complete one row per project", "section a", "section b", "section c",
    "section d", "basic project information", "project development status",
    "project title", "project name",
)


def _is_junk_project_row(name: str, row: tuple, col_map: dict) -> bool:
    """A row is scaffolding (not a project) if its name is boilerplate, or it
    has neither a sector nor a country (description/blank rows)."""
    nl = name.strip().lower()
    if any(marker in nl for marker in _JUNK_NAME_MARKERS):
        return True
    has_sector = bool(_cell(row, col_map, "pillar"))
    has_country = bool(_cell(row, col_map, "lead_country"))
    return not (has_sector or has_country)


def _select_pipeline_sheet(wb):
    """Pick the worksheet most likely to hold the project rows: a preferred tab
    name first, otherwise the sheet whose best header row scores highest."""
    for sheet_name in ("Investment Opportunities", "Projects", "Pipeline"):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
    best_ws, best_score = None, -1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _, col_map = _find_header_row(ws)
        score = len(col_map) if col_map else 0
        if score > best_score:
            best_ws, best_score = ws, score
    return best_ws or wb.active


def parse_pipeline_workbook(wb, fallback_pillar: str):
    """Pure parser: turn a pipeline workbook into a list of Project-kwarg dicts.

    `fallback_pillar` (a TWGPillar value string) is used when a row's Sector is
    blank or unrecognized — this is the importing TWG's pillar, so an Energy
    sheet imported into the Energy TWG never silently becomes Digital.

    Returns (project_dicts, skipped_count, errors).
    """
    ws = _select_pipeline_sheet(wb)
    header_row_idx, col_map = _find_header_row(ws)
    if header_row_idx is None:
        # Fall back to scanning every sheet for a usable header.
        for sheet_name in wb.sheetnames:
            candidate = wb[sheet_name]
            header_row_idx, col_map = _find_header_row(candidate)
            if header_row_idx is not None:
                ws = candidate
                break
    if header_row_idx is None:
        return [], 0, ["Could not find a header row (need Sector + at least two other known columns)."]

    projects: list[dict] = []
    skipped = 0
    errors: list[str] = []

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[header_row_idx:]  # header_row_idx is 1-based -> slice skips the header row

    for row_offset, row in enumerate(data_rows, start=header_row_idx + 2):
        name = _cell(row, col_map, "name")
        if not name or _is_junk_project_row(name, row, col_map):
            skipped += 1
            continue
        try:
            matched = _match_pillar(_cell(row, col_map, "pillar"))
            pillar = matched.value if matched else fallback_pillar

            raw_status = _cell(row, col_map, "status")
            status = _map_status(raw_status) if raw_status else ProjectStatus.DRAFT

            raw_investment = _cell(row, col_map, "investment_size")
            investment_size = _parse_investment(raw_investment) if raw_investment else 0.0

            is_cross_border = "cross" in _cell(row, col_map, "is_cross_border").lower()

            def _get(field: str) -> Optional[str]:
                v = _cell(row, col_map, field)
                return v if v and v.upper() not in ("TBC", "N/A", "NA", "NONE", "-") else None

            projects.append({
                "name": name,
                "description": "",
                "investment_size": investment_size or 0,
                "currency": "USD",
                "status": status,
                "pillar": pillar,
                "lead_country": _cell(row, col_map, "lead_country") or None,
                "subsector": _get("subsector"),
                "project_sponsor": _get("project_sponsor"),
                "is_cross_border": is_cross_border,
                "key_contact_name": _get("key_contact_name"),
                "key_contact_email": _get("key_contact_email"),
                "technical_studies": _get("technical_studies"),
                "permits_licences": _get("permits_licences"),
                "land_status": _get("land_status"),
                "financing_structure": _get("financing_structure"),
                "investment_stage_label": _get("investment_stage_label"),
                "revenue_model": _get("revenue_model"),
                "macroeconomic_roi": _get("macroeconomic_roi"),
                "climate_impact": _get("climate_impact"),
                "esg_compliance": None,
                "ghg_avoided_target": _get("ghg_avoided_target"),
                "jobs_construction": _get("jobs_construction"),
                "jobs_om": _get("jobs_om"),
                "electricity_connections": _get("electricity_connections"),
                "digital_connections": _get("digital_connections"),
                "smallholder_farmers_reached": _get("smallholder_farmers_reached"),
                "submitted_by": _get("submitted_by"),
            })
        except Exception as exc:
            errors.append(f"Row {row_offset}: {exc}")
            skipped += 1

    return projects, skipped, errors


@router.post("/import-excel")
async def import_projects_from_excel(
    file: UploadFile = File(...),
    twg_id: str = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator),
):
    """
    Bulk-import projects from an .xlsx file.
    Skips instruction rows before the header; maps columns by partial header name.
    Returns counts of imported, skipped, and per-row errors.
    """
    import openpyxl

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        parsed_twg_id = uuid.UUID(twg_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="twg_id must be a valid UUID.")

    raw_bytes = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot parse Excel file: {exc}")

    # Resolve the importing TWG's pillar — used as the fallback when a row's
    # Sector is blank/unrecognized, so projects land in the TWG they were
    # imported into rather than silently defaulting to Digital.
    twg = (await db.execute(select(TWG).where(TWG.id == parsed_twg_id))).scalar_one_or_none()
    if twg is None:
        raise HTTPException(status_code=404, detail="TWG not found.")
    fallback_pillar = twg.pillar.value if hasattr(twg.pillar, "value") else str(twg.pillar)

    project_dicts, skipped, errors = parse_pipeline_workbook(wb, fallback_pillar)
    if not project_dicts and not errors:
        raise HTTPException(
            status_code=422,
            detail="No project rows found. Check the sheet has a header row with Sector and Country columns.",
        )

    imported = 0
    for d in project_dicts:
        db.add(Project(id=uuid.uuid4(), twg_id=parsed_twg_id, **d))
        imported += 1

    try:
        await db.commit()
    except Exception as exc:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Database commit failed: {exc}")

    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.get("/{project_id}/buyer-matches", response_model=List[BuyerMatchRead])
async def get_buyer_matches(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = get_buyer_matching_service(db)
    matches = await service.get_matches_for_project(project_id)
    return [
        BuyerMatchRead(
            match_id=m["match_id"],
            buyer=BuyerRead.model_validate(m["buyer"]),
            score=m["score"],
            status=m["status"],
            match_rationale=m["match_rationale"],
        )
        for m in matches
    ]


@router.post("/{project_id}/buyer-match", response_model=dict)
async def trigger_buyer_matching(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator)
):
    service = get_buyer_matching_service(db)
    result = await service.match_buyers(project_id)
    return result


@router.get("/{project_id}/dfi-matches", response_model=List[DFIMatchRead])
async def get_dfi_matches(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retrieve scored DFI window matches for a project."""
    svc = get_dfi_matching_service(db)
    return await svc.get_matches_for_project(project_id)


@router.post("/{project_id}/dfi-match", response_model=dict)
async def trigger_dfi_matching(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Run the DFI matching engine for a project."""
    svc = get_dfi_matching_service(db)
    result = await svc.match_dfi_windows(project_id)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


@router.post("/{project_id}/financing-memo", response_model=FinancingMemoResponse)
async def get_financing_memo(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate a blended finance structuring memo for a project (LLM-powered)."""
    svc = get_dfi_matching_service(db)
    result = await svc.generate_financing_memo(project_id)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


@router.get("/{project_id}/incubation-checklist", response_model=IncubationChecklistRead)
async def get_incubation_checklist(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return the six-item Incubation document checklist for a project.

    Each slot ticks if a Document exists whose document_type matches the
    canonical code (e.g. 'FEASIBILITY') or any of its legacy aliases. Most
    recently uploaded match wins when there are several.

    Access is restricted to roles in INCUBATION_VISIBLE_ROLES.
    """
    if current_user.role not in INCUBATION_VISIBLE_ROLES:
        raise HTTPException(status_code=403, detail="Not authorised for incubation data")

    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if project.status != ProjectStatus.INCUBATION:
        raise HTTPException(
            status_code=400,
            detail="Incubation checklist is only available for projects in INCUBATION status",
        )

    docs_result = await db.execute(
        select(Document)
        .where(Document.project_id == project_id)
        .order_by(desc(Document.created_at))
    )
    docs = docs_result.scalars().all()

    # Bucket the most-recent document per canonical code.
    by_code: dict[str, Document] = {}
    for doc in docs:
        code = canonical_code_for(doc.document_type)
        if code and code not in by_code:
            by_code[code] = doc

    items: list[IncubationChecklistItem] = []
    for code, label in INCUBATION_CHECKLIST_ITEMS:
        matched = by_code.get(code)
        items.append(IncubationChecklistItem(
            code=code,
            label=label,
            completed=matched is not None,
            document_id=matched.id if matched else None,
        ))

    completed = sum(1 for i in items if i.completed)
    return IncubationChecklistRead(
        items=items,
        completed_count=completed,
        total_count=len(items),
    )


# ─────────────────────────────────────────────────────────────────────────────
# R8 — Geospatial site analysis (STUB; deterministic synthetic data)
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/{project_id}/analyse-site", response_model=ProjectGeospatialRead)
async def analyse_site(
    project_id: uuid.UUID,
    force: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator),
):
    """Run geospatial analysis on a project's `site_lat`/`site_lon`.

    A previously-analysed row within 30 days for the same coords is reused
    unless `?force=true` is passed. Coords change → cache auto-invalidates."""
    svc = get_geospatial_service(db)
    result = await svc.analyse_project(project_id, force=force)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result.get("message") or result["error"])
    return result


@router.get("/{project_id}/site-analysis", response_model=ProjectGeospatialRead)
async def get_site_analysis(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return the most-recent cached geo analysis for a project. 404 if never
    analysed."""
    result = await db.execute(
        select(ProjectGeospatialData).where(ProjectGeospatialData.project_id == project_id)
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="No site analysis yet for this project")
    return row


@router.post("/{project_id}/scout-coordinates")
async def scout_coordinates(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_facilitator),
):
    """Ask the LLM to infer plausible GPS coordinates for a project's site
    from its textual metadata. Returns a suggestion the facilitator must
    confirm — the coordinates are NOT persisted by this endpoint.

    Response shape:
        {lat, lon, place_name, confidence, reasoning, project_id}"""
    svc = get_coordinate_scout_service(db)
    result = await svc.scout(project_id)
    if "error" in result:
        msg = result.get("message") or result["error"]
        raise HTTPException(status_code=400, detail=msg)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# R9 — Post-commitment impact monitoring
# ─────────────────────────────────────────────────────────────────────────────

_COMMITTED_STATUSES = {ProjectStatus.COMMITTED, ProjectStatus.IMPLEMENTED}


def _parse_first_int(text: Optional[str]) -> Optional[int]:
    """Pull the first integer out of free-form text (e.g. '2500 jobs over 3 years' → 2500)."""
    if not text:
        return None
    m = re.search(r"\d+", str(text))
    return int(m.group(0)) if m else None


def _parse_int_sum(text: Optional[str]) -> int:
    """Sum every integer found in text. Mirrors the pattern in _compute_waiis_sub_scores."""
    if not text:
        return 0
    nums = re.findall(r"\d+", str(text))
    return sum(int(n) for n in nums) if nums else 0


def _parse_first_float(text: Optional[str]) -> Optional[float]:
    if not text:
        return None
    m = re.search(r"\d+(?:\.\d+)?", str(text))
    return float(m.group(0)) if m else None


@router.post("/{project_id}/impact-log", response_model=ImpactLogEntryRead, status_code=201)
async def create_impact_log_entry(
    project_id: uuid.UUID,
    payload: ImpactLogEntryCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Log a quarterly impact data point for a COMMITTED or IMPLEMENTED project."""
    proj_res = await db.execute(select(Project).where(Project.id == project_id))
    project = proj_res.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if project.status not in _COMMITTED_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Project must be in COMMITTED or IMPLEMENTED status to log impact data; "
                f"current status: {project.status.value}"
            ),
        )

    entry = ImpactLogEntry(
        project_id=project_id,
        period_label=payload.period_label,
        period_start=payload.period_start,
        period_end=payload.period_end,
        jobs_created=payload.jobs_created,
        ghg_avoided_tco2=payload.ghg_avoided_tco2,
        smallholders_reached=payload.smallholders_reached,
        women_jobs_actual=payload.women_jobs_actual,
        youth_jobs_actual=payload.youth_jobs_actual,
        investment_deployed_usd=payload.investment_deployed_usd,
        notes=payload.notes,
        logged_by_id=current_user.id,
    )
    db.add(entry)
    await db.commit()
    await db.refresh(entry)
    return entry


@router.get("/{project_id}/impact-log", response_model=List[ImpactLogEntryRead])
async def list_impact_log_entries(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """All impact log entries for a project, most recent period first."""
    result = await db.execute(
        select(ImpactLogEntry)
        .where(ImpactLogEntry.project_id == project_id)
        .order_by(ImpactLogEntry.period_start.desc())
    )
    return result.scalars().all()


@router.get("/{project_id}/impact-log/summary", response_model=ImpactSummaryRead)
async def get_impact_summary(
    project_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Cumulative actuals plus target values parsed from the project's text fields."""
    proj_res = await db.execute(select(Project).where(Project.id == project_id))
    project = proj_res.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    entries_res = await db.execute(
        select(ImpactLogEntry).where(ImpactLogEntry.project_id == project_id)
    )
    entries = entries_res.scalars().all()

    return ImpactSummaryRead(
        project_id=project_id,
        target_jobs=_parse_int_sum(f"{project.jobs_construction or ''} {project.jobs_om or ''}") or None,
        target_ghg_tco2=_parse_first_float(project.ghg_avoided_target),
        target_smallholders=_parse_first_int(project.smallholder_farmers_reached),
        target_investment_usd=project.investment_size,
        actual_jobs=sum(e.jobs_created or 0 for e in entries),
        actual_ghg_tco2=sum(e.ghg_avoided_tco2 or 0.0 for e in entries),
        actual_smallholders=sum(e.smallholders_reached or 0 for e in entries),
        actual_women_jobs=sum(e.women_jobs_actual or 0 for e in entries),
        actual_youth_jobs=sum(e.youth_jobs_actual or 0 for e in entries),
        actual_investment_deployed=sum((e.investment_deployed_usd or Decimal("0") for e in entries), Decimal("0")),
        entry_count=len(entries),
    )


@router.delete("/{project_id}/impact-log/{entry_id}")
async def delete_impact_log_entry(
    project_id: uuid.UUID,
    entry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    result = await db.execute(
        select(ImpactLogEntry).where(
            ImpactLogEntry.id == entry_id,
            ImpactLogEntry.project_id == project_id,
        )
    )
    entry = result.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    await db.delete(entry)
    await db.commit()
    return {"deleted": str(entry_id)}

